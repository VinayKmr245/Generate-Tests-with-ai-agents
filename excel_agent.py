"""
Agent 4 — ExcelAgent
Responsibility:
  • write_workbook   — create a new .xlsx from ctx.test_cases
  • read_test_cases  — read test cases back from an existing workbook
  • update_results   — write execution results (pass/fail/error/duration/
                       timestamp) into every relevant sheet AND create a
                       dedicated "Execution Report" summary sheet

Sheet layout after a full run
─────────────────────────────
  Summary            — static run metadata
  Execution Report   — per-run pass/fail dashboard (created/refreshed on update)
  Test Cases         — all rows, all columns, results filled in
  <Module>...        — one sheet per module
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import re
from datetime import datetime
from pathlib import Path

from config import COL, COL_WIDTHS, COLOUR, HEADERS, OUTPUT_DIR
from logger import log
from models import AgentContext, AutomationResult, TestCase
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAME = "ExcelAgent"

# ─────────────────────────────────────────────────────────────────────────────
# Style constants
# ─────────────────────────────────────────────────────────────────────────────

THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
MED_BORDER = Border(
    left=Side(style="medium"), right=Side(style="medium"),
    top=Side(style="medium"),  bottom=Side(style="medium"),
)

FONT_BASE   = Font(name="Arial", size=10)
FONT_HEADER = Font(name="Arial", size=11, bold=True, color="FFFFFF")
FONT_BOLD   = Font(name="Arial", size=10, bold=True)
FONT_TITLE  = Font(name="Arial", size=14, bold=True, color=COLOUR["header_bg"])
FONT_PASS   = Font(name="Arial", size=10, bold=True, color="375623")
FONT_FAIL   = Font(name="Arial", size=10, bold=True, color="9C0006")
FONT_ERROR  = Font(name="Arial", size=10, bold=True, color="7D6608")

FILL_HEADER  = PatternFill("solid", fgColor=COLOUR["header_bg"])
FILL_ALT     = PatternFill("solid", fgColor=COLOUR["alt_row"])
FILL_PASS    = PatternFill("solid", fgColor=COLOUR["pass_bg"])
FILL_FAIL    = PatternFill("solid", fgColor=COLOUR["fail_bg"])
FILL_ERR     = PatternFill("solid", fgColor=COLOUR["error_bg"])
FILL_SECTION = PatternFill("solid", fgColor="D9E1F2")

PRI_FILL = {
    "High":   PatternFill("solid", fgColor=COLOUR["high_pri"]),
    "Medium": PatternFill("solid", fgColor=COLOUR["med_pri"]),
    "Low":    PatternFill("solid", fgColor=COLOUR["low_pri"]),
}
RESULT_FILL = {
    "Pass":  FILL_PASS,
    "Fail":  FILL_FAIL,
    "Error": FILL_ERR,
}
RESULT_FONT = {
    "Pass":  FONT_PASS,
    "Fail":  FONT_FAIL,
    "Error": FONT_ERROR,
}


# ─────────────────────────────────────────────────────────────────────────────
# Cell helpers
# ─────────────────────────────────────────────────────────────────────────────

def _hdr(cell, text: str):
    cell.value     = text
    cell.font      = FONT_HEADER
    cell.fill      = FILL_HEADER
    cell.border    = THIN
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _data(cell, value, alt=False, wrap=True, center=False):
    cell.value     = value
    cell.font      = FONT_BASE
    cell.border    = THIN
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="top",
        wrap_text=wrap,
    )
    if alt:
        cell.fill = FILL_ALT


def _apply_col_widths(ws):
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_header_row(ws):
    for i, h in enumerate(HEADERS, 1):
        _hdr(ws.cell(row=1, column=i), h)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def _write_tc_row(ws, row: int, tc: TestCase):
    alt = row % 2 == 0
    vals = [
        tc.test_case_id, tc.module, tc.test_case_title, tc.description,
        tc.preconditions, tc.test_steps, tc.expected_result, tc.priority,
        tc.test_type, tc.status, tc.auto_script, tc.auto_result,
        tc.auto_error, tc.executed_at, tc.comments,
    ]
    for col_idx, val in enumerate(vals, 1):
        _data(ws.cell(row=row, column=col_idx), val, alt=alt)

    # Priority cell
    pri_cell            = ws.cell(row=row, column=COL["priority"])
    pri_cell.fill       = PRI_FILL.get(tc.priority.strip().title(), PatternFill())
    pri_cell.font       = FONT_BOLD
    pri_cell.alignment  = Alignment(horizontal="center", vertical="top")

    # Result cell
    res_label = tc.auto_result.strip().title()
    if res_label:
        res_cell           = ws.cell(row=row, column=COL["auto_result"])
        res_cell.fill      = RESULT_FILL.get(res_label, PatternFill())
        res_cell.font      = RESULT_FONT.get(res_label, FONT_BOLD)
        res_cell.alignment = Alignment(horizontal="center", vertical="top")

    ws.row_dimensions[row].height = 60


# ─────────────────────────────────────────────────────────────────────────────
# Execution Report sheet builder
# ─────────────────────────────────────────────────────────────────────────────

_ER_COLS = [
    ("Test Case ID", 14), ("Title", 38), ("Module", 20),
    ("Priority", 12),     ("Type", 18),  ("Result", 12),
    ("Duration (ms)", 15), ("Executed At", 22), ("Error Summary", 55),
]

def _build_execution_report(wb: Workbook, results: list[AutomationResult],
                             test_cases: list[TestCase], run_ts: str):
    """
    Create (or replace) the 'Execution Report' sheet with:
      - Run statistics banner
      - Per-test result table with duration and error summary
    """
    # Remove existing sheet if present
    if "Execution Report" in wb.sheetnames:
        del wb["Execution Report"]

    # Insert as second sheet
    er = wb.create_sheet("Execution Report", 1)

    # Build lookup maps
    result_map  = {r.test_case_id: r for r in results}
    tc_map      = {tc.test_case_id: tc for tc in test_cases}

    total   = len(results)
    passed  = sum(1 for r in results if r.passed)
    failed  = total - passed
    errors  = sum(1 for r in results if not r.passed and "Error" in r.error_message[:20])
    avg_ms  = int(sum(r.duration_ms for r in results) / max(total, 1))
    pct     = f"{(passed/max(total,1)*100):.1f}%"

    # ── Banner (rows 1-8) ────────────────────────────────────────────────────
    er.merge_cells("A1:I1")
    title_cell       = er["A1"]
    title_cell.value = "Execution Report"
    title_cell.font  = FONT_TITLE
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    er.row_dimensions[1].height = 36

    stats = [
        ("Run Timestamp",  run_ts),
        ("Total Tests",    total),
        ("✅  Passed",      passed),
        ("❌  Failed",      failed),
        ("Pass Rate",       pct),
        ("Avg Duration",   f"{avg_ms} ms"),
    ]
    for r_off, (label, value) in enumerate(stats, start=2):
        lbl_cell       = er.cell(row=r_off, column=1, value=label)
        lbl_cell.font  = FONT_BOLD
        lbl_cell.fill  = FILL_SECTION
        lbl_cell.border = THIN
        lbl_cell.alignment = Alignment(horizontal="left", vertical="center")

        val_cell       = er.cell(row=r_off, column=2, value=value)
        val_cell.font  = FONT_BASE
        val_cell.border = THIN
        val_cell.alignment = Alignment(horizontal="left", vertical="center")

        # Colour pass/fail counts
        if label == "✅  Passed":
            val_cell.fill = FILL_PASS
            val_cell.font = FONT_PASS
        elif label == "❌  Failed":
            val_cell.fill = FILL_FAIL
            val_cell.font = FONT_FAIL

        er.row_dimensions[r_off].height = 20

    # ── Results table header (row 9) ─────────────────────────────────────────
    TABLE_START = 9
    er.row_dimensions[TABLE_START].height = 28
    for col_idx, (hdr_text, width) in enumerate(_ER_COLS, 1):
        _hdr(er.cell(row=TABLE_START, column=col_idx), hdr_text)
        er.column_dimensions[get_column_letter(col_idx)].width = width

    er.freeze_panes = f"A{TABLE_START + 1}"

    # ── Results rows ──────────────────────────────────────────────────────────
    for row_off, r in enumerate(results, start=TABLE_START + 1):
        alt = row_off % 2 == 0
        tc  = tc_map.get(r.test_case_id)

        result_label = "Pass" if r.passed else (
            "Error" if r.error_message and r.error_message.startswith("Execution error")
            else "Fail"
        )

        # One-line error summary (last non-blank line of traceback)
        error_lines   = [l.strip() for l in r.error_message.splitlines() if l.strip()]
        error_summary = error_lines[-1][:200] if error_lines else ""

        row_vals = [
            r.test_case_id,
            tc.test_case_title if tc else "",
            tc.module          if tc else "",
            tc.priority        if tc else "",
            tc.test_type       if tc else "",
            result_label,
            r.duration_ms,
            r.executed_at,
            error_summary,
        ]

        for col_idx, val in enumerate(row_vals, 1):
            cell = er.cell(row=row_off, column=col_idx)
            _data(cell, val, alt=alt, wrap=True,
                  center=(col_idx in (1, 4, 5, 6, 7)))

        # Colour the Result cell
        res_cell      = er.cell(row=row_off, column=6)
        res_cell.fill = RESULT_FILL.get(result_label, PatternFill())
        res_cell.font = RESULT_FONT.get(result_label, FONT_BOLD)

        er.row_dimensions[row_off].height = 22

    # Column A (ID) width
    er.column_dimensions["A"].width = 14


# ─────────────────────────────────────────────────────────────────────────────
# write_workbook
# ─────────────────────────────────────────────────────────────────────────────

def write_workbook(ctx: AgentContext) -> str:
    test_cases = ctx.test_cases
    url        = ctx.url
    page_title = ctx.page_analyses[-1].page_title if ctx.page_analyses else url

    wb = Workbook()
    wb.remove(wb.active)

    # ── Summary sheet ──────────────────────────────────────────────────────
    ss = wb.create_sheet("Summary")
    ss.column_dimensions["A"].width = 28
    ss.column_dimensions["B"].width = 60

    ss["A1"] = "Test Report Summary"
    ss["A1"].font = FONT_TITLE
    ss.merge_cells("A1:B1")
    ss["A1"].alignment = Alignment(horizontal="center")

    rows = [
        ("Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("URL Tested", url),
        ("Page Title", page_title),
        ("Total Test Cases", len(test_cases)),
        ("High Priority",   sum(1 for t in test_cases if t.priority.title() == "High")),
        ("Medium Priority", sum(1 for t in test_cases if t.priority.title() == "Medium")),
        ("Low Priority",    sum(1 for t in test_cases if t.priority.title() == "Low")),
    ]
    for r_idx, (lbl, val) in enumerate(rows, start=3):
        ss.cell(row=r_idx, column=1, value=lbl).font = FONT_BOLD
        ss.cell(row=r_idx, column=2, value=val).font = FONT_BASE

    # ── All Test Cases sheet ───────────────────────────────────────────────
    ts_ws = wb.create_sheet("Test Cases")
    _write_header_row(ts_ws)
    _apply_col_widths(ts_ws)
    for i, tc in enumerate(test_cases, start=2):
        _write_tc_row(ts_ws, i, tc)

    # ── Per-module sheets ──────────────────────────────────────────────────
    modules: dict[str, list[TestCase]] = {}
    for tc in test_cases:
        modules.setdefault(tc.module, []).append(tc)

    for mod_name, mod_cases in modules.items():
        safe = re.sub(r"[\\/*?:\[\]]", "_", mod_name)[:31]
        ms   = wb.create_sheet(safe)
        _write_header_row(ms)
        _apply_col_widths(ms)
        for i, tc in enumerate(mod_cases, start=2):
            _write_tc_row(ms, i, tc)

    # ── Save ───────────────────────────────────────────────────────────────
    stamp      = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_title = re.sub(r"[^\w\-]", "_", page_title)[:40]
    path       = OUTPUT_DIR / f"test_cases_{safe_title}_{stamp}.xlsx"
    wb.save(str(path))

    ctx.excel_path = str(path)
    log("excel", NAME, f"Workbook written → {path}")
    return str(path)


# ─────────────────────────────────────────────────────────────────────────────
# read_test_cases
# ─────────────────────────────────────────────────────────────────────────────

def read_test_cases(excel_path: str) -> list[TestCase]:
    wb = load_workbook(excel_path, data_only=True)
    if "Test Cases" not in wb.sheetnames:
        log("error", NAME, "Sheet 'Test Cases' not found")
        return []

    ws         = wb["Test Cases"]
    test_cases = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        tc = TestCase(
            test_case_id    = str(row[COL["id"] - 1]            or ""),
            module          = str(row[COL["module"] - 1]        or ""),
            test_case_title = str(row[COL["title"] - 1]         or ""),
            description     = str(row[COL["description"] - 1]   or ""),
            preconditions   = str(row[COL["preconditions"] - 1] or ""),
            test_steps      = str(row[COL["steps"] - 1]         or ""),
            expected_result = str(row[COL["expected"] - 1]      or ""),
            priority        = str(row[COL["priority"] - 1]      or "Medium"),
            test_type       = str(row[COL["test_type"] - 1]     or "Functional"),
            status          = str(row[COL["status"] - 1]        or "Not Executed"),
            auto_script     = str(row[COL["auto_script"] - 1]   or ""),
            auto_result     = str(row[COL["auto_result"] - 1]   or ""),
            auto_error      = str(row[COL["auto_error"] - 1]    or ""),
            executed_at     = str(row[COL["executed_at"] - 1]   or ""),
            comments        = str(row[COL["comments"] - 1]      or ""),
        )
        test_cases.append(tc)

    log("excel", NAME, f"Read {len(test_cases)} test cases from {excel_path}")
    return test_cases


# ─────────────────────────────────────────────────────────────────────────────
# update_results  ← the key method
# ─────────────────────────────────────────────────────────────────────────────

def update_results(excel_path: str, results: list[AutomationResult],
                   test_cases: list[TestCase] | None = None):
    """
    1. Update every data sheet (Test Cases + module sheets) in-place:
       auto_result, auto_error, executed_at, duration, status
    2. Build/refresh the 'Execution Report' sheet
    3. Refresh execution stats on the Summary sheet
    """
    result_map = {r.test_case_id: r for r in results}
    run_ts     = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    wb = load_workbook(excel_path)

    # ── Step 1: update data rows in all test-case sheets ─────────────────────
    for sheet_name in wb.sheetnames:
        if sheet_name in ("Summary", "Execution Report"):
            continue
        ws = wb[sheet_name]

        for row_idx in range(2, ws.max_row + 1):
            tc_id = ws.cell(row=row_idx, column=COL["id"]).value
            if not tc_id or str(tc_id) not in result_map:
                continue

            r   = result_map[str(tc_id)]
            alt = row_idx % 2 == 0

            result_label = "Pass" if r.passed else (
                "Error" if r.error_message.startswith("Execution error")
                else "Fail"
            )

            # Helper: write a plain data cell
            def _w(col_key, value, center=False):
                cell = ws.cell(row=row_idx, column=COL[col_key])
                cell.value     = value
                cell.font      = FONT_BASE
                cell.border    = THIN
                cell.alignment = Alignment(
                    horizontal="center" if center else "left",
                    vertical="top", wrap_text=True,
                )
                if alt:
                    cell.fill = FILL_ALT

            _w("auto_script",  r.script)
            _w("auto_error",   r.error_message)
            _w("executed_at",  r.executed_at,  center=True)

            # Duration — store in the "comments" column if no dedicated col,
            # or append to executed_at cell (we embed it as " (Xms)")
            exec_cell = ws.cell(row=row_idx, column=COL["executed_at"])
            exec_cell.value = f"{r.executed_at}  ({r.duration_ms} ms)"

            # Auto result — coloured
            res_cell           = ws.cell(row=row_idx, column=COL["auto_result"])
            res_cell.value     = result_label
            res_cell.font      = RESULT_FONT.get(result_label, FONT_BOLD)
            res_cell.fill      = RESULT_FILL.get(result_label, PatternFill())
            res_cell.border    = THIN
            res_cell.alignment = Alignment(horizontal="center", vertical="top")

            # Status
            status_cell        = ws.cell(row=row_idx, column=COL["status"])
            status_cell.value  = "Executed"
            status_cell.font   = FONT_BASE
            status_cell.border = THIN
            status_cell.alignment = Alignment(horizontal="center", vertical="top")

    # ── Step 2: Execution Report sheet ───────────────────────────────────────
    # We need the full TestCase list; try to rebuild from the workbook if not passed
    if test_cases is None:
        test_cases = read_test_cases(excel_path)   # re-reads from the already-open wb? No — safe, uses data_only

    _build_execution_report(wb, results, test_cases, run_ts)

    # ── Step 3: Refresh Summary stats ────────────────────────────────────────
    if "Summary" in wb.sheetnames:
        ss     = wb["Summary"]
        total  = len(results)
        passed = sum(1 for r in results if r.passed)
        failed = total - passed
        avg_ms = int(sum(r.duration_ms for r in results) / max(total, 1))

        exec_stats = [
            ("── Execution Results ──",  ""),
            ("Last Run At",              run_ts),
            ("Tests Executed",           total),
            ("Passed",                   passed),
            ("Failed",                   failed),
            ("Pass Rate",                f"{passed/max(total,1)*100:.1f}%"),
            ("Avg Duration",             f"{avg_ms} ms"),
        ]
        start_row = 12   # leave rows 3-11 for static metadata
        for i, (lbl, val) in enumerate(exec_stats):
            lbl_cell = ss.cell(row=start_row + i, column=1, value=lbl)
            val_cell = ss.cell(row=start_row + i, column=2, value=val)
            lbl_cell.font = FONT_BOLD
            val_cell.font = FONT_BASE

            if lbl == "Passed":
                val_cell.fill = FILL_PASS
                val_cell.font = FONT_PASS
            elif lbl == "Failed":
                val_cell.fill = FILL_FAIL
                val_cell.font = FONT_FAIL

    wb.save(excel_path)
    log("excel", NAME,
        f"Results written — "
        f"{sum(1 for r in results if r.passed)} passed, "
        f"{sum(1 for r in results if not r.passed)} failed → {excel_path}")


# ─────────────────────────────────────────────────────────────────────────────
# Agent entry-point
# ─────────────────────────────────────────────────────────────────────────────

def run(ctx: AgentContext, mode: str = "write") -> str:
    if mode == "write":
        return write_workbook(ctx)
    elif mode == "update":
        update_results(ctx.excel_path, ctx.automation_results, ctx.test_cases)
        return ctx.excel_path
    else:
        raise ValueError(f"Unknown ExcelAgent mode: {mode}")